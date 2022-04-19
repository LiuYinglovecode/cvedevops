#!/usr/bin/env python3

from openpyxl import Workbook
import openpyxl
import time

#book = Workbook()
# write
#sheet['A1'] = 56
#sheet['A2'] = 43

#now = time.strftime("%x")
#sheet['A3'] = now

#append
#rows = (
#    (88, 46, 57),
#    (89, 38, 12),
#    (23, 59, 78),
#    (56, 21, 98),
#    (24, 18, 43),
#    (34, 15, 67)
#)

#for row in rows:
#    sheet.append(row)

# Read
book = openpyxl.load_workbook('march.xlsx')
sheet = book.active
ip = sheet['B2']
cve = sheet['G2']
status = sheet['Q2']
alert =sheet['F2']

booktwo = Workbook()
sheet2 = booktwo.active

sheet2['A2'] = ip.value
sheet2['G2'] = cve.value
sheet2['Q2'] = status.value
sheet2['F2'] = alert.value

booktwo.save("read.xlsx")
